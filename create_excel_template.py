import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

HEADERS_PROCESS = [
    "test_case_id","process_name","description","process_type","transfer_mode",
    "source_datasource","source_schema","source_table",
    "target_datasource","target_schema","target_table",
    "truncate_target","create_table","schedule_enabled","cron_expression",
    "expected_status","expected_row_count",
]
HEADERS_RELEASE = HEADERS_PROCESS + [
    "agent_name","host","port","agent_type","ds_name","db_type","database","username","password","schema"
]
SHEETS = {
    "BulkTransfer_DB2_MSSQL": HEADERS_PROCESS,
    "BulkTransfer_DB2_Postgres": HEADERS_PROCESS,
    "BulkTransfer_VSAM_MSSQL": HEADERS_PROCESS,
    "Release_30_05_25": HEADERS_RELEASE,
    "Release_27_06_25": HEADERS_PROCESS,
}
SAMPLE = {h: "" for h in HEADERS_RELEASE}
SAMPLE.update({"test_case_id":"TC001","process_name":"Sample_Process","process_type":"BULK_TRANSFER",
               "transfer_mode":"FULL","expected_status":"COMPLETED","truncate_target":"true","create_table":"false"})

def style(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(str(cell.value or ""))+4, 18)

def main():
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "RDRS.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for name, headers in SHEETS.items():
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        ws.append([SAMPLE.get(h,"") for h in headers])
        style(ws); ws.freeze_panes = "A2"
    wb.save(out)
    print(f"Excel template: {out}")

if __name__ == "__main__":
    main()
