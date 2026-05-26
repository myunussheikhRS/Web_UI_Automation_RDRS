import os, json, textwrap

BASE = r"C:\Users\msheikh\RDRS_Web_Automation"

def w(rel, content):
    p = os.path.join(BASE, rel.replace("/", os.sep))
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        f.write(textwrap.dedent(content).lstrip("\n"))
    print("OK:", rel)

def wj(rel, data):
    p = os.path.join(BASE, rel.replace("/", os.sep))
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print("OK:", rel)

# remaining locators
wj("locators/agent_page.json", {
  "btn_add_agent":{"type":"css","value":"button#addAgent, button.add-agent"},
  "btn_edit_agent":{"type":"css","value":"button.edit-agent, a.edit-agent, .agent-action-edit"},
  "btn_delete_agent":{"type":"css","value":"button.delete-agent, .agent-action-delete"},
  "btn_save":{"type":"css","value":"button#saveAgent, button[type='submit'].save-btn"},
  "btn_cancel":{"type":"css","value":"button#cancelAgent, button.cancel-btn"},
  "btn_confirm_delete":{"type":"css","value":"button#confirmDelete, .modal button.btn-danger"},
  "input_agent_name":{"type":"css","value":"input#agentName, input[name='agentName']"},
  "input_host":{"type":"css","value":"input#host, input[name='host']"},
  "input_port":{"type":"css","value":"input#port, input[name='port']"},
  "input_description":{"type":"css","value":"textarea#description, input#description"},
  "select_agent_type":{"type":"css","value":"select#agentType, select[name='agentType']"},
  "agent_table":{"type":"css","value":"table#agentTable, table.agent-list"},
  "agent_table_rows":{"type":"css","value":"table#agentTable tbody tr, table.agent-list tbody tr"},
  "search_agent":{"type":"css","value":"input#searchAgent, input.search-agent"},
  "page_title":{"type":"css","value":"h1.page-title, .page-header h1"},
  "no_records_msg":{"type":"css","value":".no-records, td.dataTables_empty"},
})

wj("locators/data_source_page.json", {
  "btn_add_datasource":{"type":"css","value":"button#addDataSource, button.add-datasource"},
  "btn_edit_datasource":{"type":"css","value":"button.edit-datasource, .ds-action-edit"},
  "btn_delete_datasource":{"type":"css","value":"button.delete-datasource, .ds-action-delete"},
  "btn_save":{"type":"css","value":"button#saveDataSource, button[type='submit'].save-btn"},
  "btn_cancel":{"type":"css","value":"button#cancelDataSource, button.cancel-btn"},
  "btn_test_connection":{"type":"css","value":"button#testConnection, button.test-connection"},
  "btn_confirm_delete":{"type":"css","value":"button#confirmDelete, .modal button.btn-danger"},
  "input_ds_name":{"type":"css","value":"input#dataSourceName, input[name='dataSourceName']"},
  "input_host":{"type":"css","value":"input#dbHost, input[name='host']"},
  "input_port":{"type":"css","value":"input#dbPort, input[name='port']"},
  "input_database":{"type":"css","value":"input#database, input[name='database']"},
  "input_username":{"type":"css","value":"input#username, input[name='username']"},
  "input_password":{"type":"css","value":"input#password, input[name='password']"},
  "input_schema":{"type":"css","value":"input#schema, input[name='schema']"},
  "select_db_type":{"type":"css","value":"select#dbType, select[name='dbType']"},
  "select_agent":{"type":"css","value":"select#agent, select[name='agent']"},
  "datasource_table":{"type":"css","value":"table#dataSourceTable, table.datasource-list"},
  "datasource_table_rows":{"type":"css","value":"table#dataSourceTable tbody tr"},
  "connection_success_msg":{"type":"css","value":".connection-success, .alert-success"},
  "connection_error_msg":{"type":"css","value":".connection-error, .alert-danger"},
  "page_title":{"type":"css","value":"h1.page-title, #datasource-page-title"},
})

wj("locators/process_definition_page.json", {
  "btn_add_process":{"type":"css","value":"button#addProcess, button.add-process"},
  "btn_edit_process":{"type":"css","value":"button.edit-process, a.edit-process, .process-action-edit"},
  "btn_delete_process":{"type":"css","value":"button.delete-process, .process-action-delete"},
  "btn_save":{"type":"css","value":"button#saveProcess, button[type='submit'].save-btn"},
  "btn_cancel":{"type":"css","value":"button#cancelProcess, button.cancel-btn"},
  "btn_confirm_delete":{"type":"css","value":"button#confirmDelete, .modal button.btn-danger"},
  "btn_next_step":{"type":"css","value":"button#nextStep, button.next-step, .wizard-next"},
  "btn_prev_step":{"type":"css","value":"button#prevStep, button.prev-step"},
  "btn_finish":{"type":"css","value":"button#finishProcess, button.finish-btn, .wizard-finish"},
  "input_process_name":{"type":"css","value":"input#processName, input[name='processName']"},
  "input_description":{"type":"css","value":"textarea#description, input#description"},
  "select_source_ds":{"type":"css","value":"select#sourceDataSource, select[name='sourceDataSource']"},
  "select_target_ds":{"type":"css","value":"select#targetDataSource, select[name='targetDataSource']"},
  "select_process_type":{"type":"css","value":"select#processType, select[name='processType']"},
  "select_transfer_mode":{"type":"css","value":"select#transferMode, select[name='transferMode']"},
  "input_source_schema":{"type":"css","value":"input#sourceSchema, input[name='sourceSchema']"},
  "input_source_table":{"type":"css","value":"input#sourceTable, input[name='sourceTable']"},
  "input_target_schema":{"type":"css","value":"input#targetSchema, input[name='targetSchema']"},
  "input_target_table":{"type":"css","value":"input#targetTable, input[name='targetTable']"},
  "checkbox_truncate_target":{"type":"css","value":"input#truncateTarget, input[name='truncateTarget']"},
  "checkbox_create_table":{"type":"css","value":"input#createTable, input[name='createTable']"},
  "process_table":{"type":"css","value":"table#processTable, table.process-list"},
  "process_table_rows":{"type":"css","value":"table#processTable tbody tr"},
  "search_process":{"type":"css","value":"input#searchProcess, input.search-process"},
  "page_title":{"type":"css","value":"h1.page-title, #process-def-page-title"},
  "tab_general":{"type":"css","value":"a[href*='general'], #tab-general"},
  "tab_mapping":{"type":"css","value":"a[href*='mapping'], #tab-mapping"},
  "tab_schedule":{"type":"css","value":"a[href*='schedule'], #tab-schedule"},
})

wj("locators/process_list_page.json", {
  "process_table":{"type":"css","value":"table#processListTable, table.process-list-grid"},
  "process_table_rows":{"type":"css","value":"table#processListTable tbody tr"},
  "btn_run_process":{"type":"css","value":"button.run-process, a.run-process, .process-action-run"},
  "btn_stop_process":{"type":"css","value":"button.stop-process, a.stop-process"},
  "btn_view_log":{"type":"css","value":"button.view-log, a.view-log, .process-action-log"},
  "btn_refresh":{"type":"css","value":"button#refreshList, button.refresh-btn"},
  "search_process":{"type":"css","value":"input#searchProcessList, input.search-process-list"},
  "filter_status":{"type":"css","value":"select#filterStatus, select[name='filterStatus']"},
  "status_badge_running":{"type":"css","value":".badge-running, span.status[data-status='RUNNING']"},
  "status_badge_completed":{"type":"css","value":".badge-completed, span.status[data-status='COMPLETED']"},
  "status_badge_failed":{"type":"css","value":".badge-failed, span.status[data-status='FAILED']"},
  "status_badge_stopped":{"type":"css","value":".badge-stopped, span.status[data-status='STOPPED']"},
  "log_modal":{"type":"css","value":"#logModal, .log-modal"},
  "log_content":{"type":"css","value":"#logContent, .log-content"},
  "btn_close_log":{"type":"css","value":"#logModal .close, .log-modal .btn-close"},
  "page_title":{"type":"css","value":"h1.page-title, #process-list-page-title"},
})

wj("locators/repository_page.json", {
  "btn_add_repository":{"type":"css","value":"button#addRepository, button.add-repository"},
  "btn_edit_repository":{"type":"css","value":"button.edit-repository, .repo-action-edit"},
  "btn_delete_repository":{"type":"css","value":"button.delete-repository, .repo-action-delete"},
  "btn_save":{"type":"css","value":"button#saveRepository, button[type='submit'].save-btn"},
  "btn_cancel":{"type":"css","value":"button#cancelRepository, button.cancel-btn"},
  "btn_test_connection":{"type":"css","value":"button#testRepoConnection"},
  "btn_confirm_delete":{"type":"css","value":"button#confirmDelete, .modal button.btn-danger"},
  "input_repo_name":{"type":"css","value":"input#repositoryName, input[name='repositoryName']"},
  "input_host":{"type":"css","value":"input#repoHost, input[name='repoHost']"},
  "input_port":{"type":"css","value":"input#repoPort, input[name='repoPort']"},
  "input_database":{"type":"css","value":"input#repoDatabase, input[name='repoDatabase']"},
  "input_username":{"type":"css","value":"input#repoUsername, input[name='repoUsername']"},
  "input_password":{"type":"css","value":"input#repoPassword, input[name='repoPassword']"},
  "select_repo_type":{"type":"css","value":"select#repoType, select[name='repoType']"},
  "repository_table":{"type":"css","value":"table#repositoryTable, table.repository-list"},
  "repository_table_rows":{"type":"css","value":"table#repositoryTable tbody tr"},
  "connection_success_msg":{"type":"css","value":".connection-success, .alert-success"},
  "page_title":{"type":"css","value":"h1.page-title, #repository-page-title"},
})

wj("locators/output_target_page.json", {
  "btn_add_output_target":{"type":"css","value":"button#addOutputTarget, button.add-output-target"},
  "btn_edit_output_target":{"type":"css","value":"button.edit-output-target, .ot-action-edit"},
  "btn_delete_output_target":{"type":"css","value":"button.delete-output-target, .ot-action-delete"},
  "btn_save":{"type":"css","value":"button#saveOutputTarget, button[type='submit'].save-btn"},
  "btn_cancel":{"type":"css","value":"button#cancelOutputTarget, button.cancel-btn"},
  "btn_test_connection":{"type":"css","value":"button#testOTConnection"},
  "btn_confirm_delete":{"type":"css","value":"button#confirmDelete, .modal button.btn-danger"},
  "input_ot_name":{"type":"css","value":"input#outputTargetName, input[name='outputTargetName']"},
  "input_host":{"type":"css","value":"input#otHost, input[name='otHost']"},
  "input_port":{"type":"css","value":"input#otPort, input[name='otPort']"},
  "input_database":{"type":"css","value":"input#otDatabase, input[name='otDatabase']"},
  "input_username":{"type":"css","value":"input#otUsername, input[name='otUsername']"},
  "input_password":{"type":"css","value":"input#otPassword, input[name='otPassword']"},
  "input_schema":{"type":"css","value":"input#otSchema, input[name='otSchema']"},
  "select_ot_type":{"type":"css","value":"select#outputTargetType, select[name='outputTargetType']"},
  "select_agent":{"type":"css","value":"select#otAgent, select[name='otAgent']"},
  "output_target_table":{"type":"css","value":"table#outputTargetTable, table.output-target-list"},
  "output_target_table_rows":{"type":"css","value":"table#outputTargetTable tbody tr"},
  "connection_success_msg":{"type":"css","value":".connection-success, .alert-success"},
  "page_title":{"type":"css","value":"h1.page-title, #output-target-page-title"},
})

print("locators done")

# ── tests/conftest.py ─────────────────────────────────────────────────────────
w("tests/__init__.py", "")
w("tests/conftest.py", """
import os, sys
import allure
import pytest
from playwright.sync_api import Browser, BrowserContext, Page, sync_playwright

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from utils.config_manager import ConfigManager
from utils.screenshot_util import ScreenshotUtil

@pytest.fixture(scope="session")
def config():
    return ConfigManager()

@pytest.fixture(scope="session")
def playwright_instance():
    with sync_playwright() as pw:
        yield pw

@pytest.fixture(scope="session")
def browser(playwright_instance, config):
    opts = {"headless": config.headless}
    name = config.browser.lower()
    if name == "chromium": br = playwright_instance.chromium.launch(**opts)
    elif name == "firefox": br = playwright_instance.firefox.launch(**opts)
    elif name == "webkit":  br = playwright_instance.webkit.launch(**opts)
    else: raise ValueError(f"Unsupported browser: {name}")
    yield br
    br.close()

@pytest.fixture(scope="function")
def context(browser):
    ctx = browser.new_context(viewport={"width":1920,"height":1080}, ignore_https_errors=True)
    yield ctx
    ctx.close()

@pytest.fixture(scope="function")
def page(context):
    pg = context.new_page()
    yield pg
    pg.close()

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    if report.when == "call" and report.failed:
        pg = item.funcargs.get("page")
        if pg:
            name = item.nodeid.replace("::", "_").replace("/", "_")
            try: ScreenshotUtil.capture_on_failure(pg, name)
            except Exception: pass

def pytest_configure(config):
    os.makedirs("allure-results", exist_ok=True)
    os.makedirs("reports/screenshots", exist_ok=True)
    cfg = ConfigManager()
    with open("allure-results/environment.properties", "w") as f:
        f.write(f"application.url={cfg.web_url}\\n")
        f.write(f"browser={cfg.browser}\\n")
        f.write(f"headless={cfg.headless}\\n")
""")

# ── test suites ───────────────────────────────────────────────────────────────
for pkg in ["bulk_transfer_db2_mssql","bulk_transfer_db2_postgres",
            "bulk_transfer_vsam_mssql","release_30_05_25","release_27_06_25"]:
    w(f"tests/{pkg}/__init__.py", "")

w("tests/bulk_transfer_db2_mssql/test_bulk_process_1.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from pages.process_list_page import ProcessListPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "BulkTransfer_DB2_MSSQL"

@allure.feature("Bulk Transfer")
@allure.story("DB2 to MSSQL")
@pytest.mark.bulk_transfer_db2_mssql
class TestBulkProcessDB2MSSQL:

    @allure.title("TC001 - Create and run bulk transfer DB2 to MSSQL")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_bulk_transfer_tc001(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC001", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_add_process().fill_process_form(
            process_name=data.process_name, process_type=data.process_type,
            source_ds=data.source_datasource, target_ds=data.target_datasource,
            source_schema=data.source_schema, source_table=data.source_table,
            target_schema=data.target_schema, target_table=data.target_table,
            transfer_mode=data.transfer_mode, truncate_target=data.truncate_target,
            create_table=data.create_table, description=data.description,
        ).click_save()
        AssertUtil.assert_true(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' visible after save")
        HomePage(page).go_to_process_list()
        pl = ProcessListPage(page)
        pl.click_run_process(data.process_name)
        pl.wait_for_process_status(data.process_name, data.expected_status, 120000)
        AssertUtil.assert_equals(pl.get_process_status(data.process_name).upper(),
            data.expected_status.upper(), "Process status")

    @allure.title("TC002 - Verify process visible in process list")
    @allure.severity(allure.severity_level.NORMAL)
    def test_bulk_transfer_tc002(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC002", SHEET))
        HomePage(page).open().go_to_process_list()
        AssertUtil.assert_true(ProcessListPage(page).is_process_present(data.process_name),
            f"Process '{data.process_name}' in list")
""")

w("tests/bulk_transfer_db2_mssql/test_bulk_process_2.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "BulkTransfer_DB2_MSSQL"

@allure.feature("Bulk Transfer")
@allure.story("DB2 to MSSQL - Extended")
@pytest.mark.bulk_transfer_db2_mssql
class TestBulkProcessDB2MSSQLExtended:

    @allure.title("TC003 - Delete bulk transfer process")
    def test_delete_process_tc003(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC003", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_delete_process(data.process_name).confirm_delete()
        AssertUtil.assert_false(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' removed")

    @allure.title("TC004 - Edit bulk transfer process")
    def test_edit_process_tc004(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC004", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_edit_process(data.process_name).enter_description(data.description).click_save()
        AssertUtil.assert_true(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' still present after edit")
""")

w("tests/bulk_transfer_db2_postgres/test_bulk_db2_postgres_1.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from pages.process_list_page import ProcessListPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "BulkTransfer_DB2_Postgres"

@allure.feature("Bulk Transfer")
@allure.story("DB2 to PostgreSQL")
@pytest.mark.bulk_transfer_db2_postgres
class TestBulkProcessDB2Postgres:

    @allure.title("TC001 - Create and run bulk transfer DB2 to PostgreSQL")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_bulk_transfer_tc001(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC001", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_add_process().fill_process_form(
            process_name=data.process_name, process_type=data.process_type,
            source_ds=data.source_datasource, target_ds=data.target_datasource,
            source_schema=data.source_schema, source_table=data.source_table,
            target_schema=data.target_schema, target_table=data.target_table,
            transfer_mode=data.transfer_mode, truncate_target=data.truncate_target,
        ).click_save()
        AssertUtil.assert_true(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' visible after save")
        HomePage(page).go_to_process_list()
        pl = ProcessListPage(page)
        pl.click_run_process(data.process_name)
        pl.wait_for_process_status(data.process_name, data.expected_status, 120000)
        AssertUtil.assert_equals(pl.get_process_status(data.process_name).upper(),
            data.expected_status.upper(), "Process status")
""")

w("tests/bulk_transfer_vsam_mssql/test_bulk_vsam_mssql_1.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from pages.process_list_page import ProcessListPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "BulkTransfer_VSAM_MSSQL"

@allure.feature("Bulk Transfer")
@allure.story("VSAM to MSSQL")
@pytest.mark.bulk_transfer_vsam_mssql
class TestBulkProcessVSAMMSSQL:

    @allure.title("TC001 - Create and run bulk transfer VSAM to MSSQL")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_bulk_transfer_tc001(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC001", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_add_process().fill_process_form(
            process_name=data.process_name, process_type=data.process_type,
            source_ds=data.source_datasource, target_ds=data.target_datasource,
            source_schema=data.source_schema, source_table=data.source_table,
            target_schema=data.target_schema, target_table=data.target_table,
            transfer_mode=data.transfer_mode, truncate_target=data.truncate_target,
        ).click_save()
        AssertUtil.assert_true(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' visible after save")
        HomePage(page).go_to_process_list()
        pl = ProcessListPage(page)
        pl.click_run_process(data.process_name)
        pl.wait_for_process_status(data.process_name, data.expected_status, 120000)
        AssertUtil.assert_equals(pl.get_process_status(data.process_name).upper(),
            data.expected_status.upper(), "Process status")
""")

w("tests/release_30_05_25/test_rdrs_1011_tc001.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.agent_page import AgentPage
from pages.data_source_page import DataSourcePage
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from pages.process_list_page import ProcessListPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "Release_30_05_25"

@allure.feature("Release 30-05-25")
@allure.story("RDRS-1011")
@pytest.mark.release_30_05_25
class TestRDRS1011:

    @allure.title("RDRS-1011 TC001 - Agent creation")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_rdrs_1011_tc001(self, page):
        raw = ExcelReader().get_test_data("RDRS_1011_TC001", SHEET)
        HomePage(page).open().go_to_agent()
        agent = AgentPage(page)
        agent.click_add_agent().fill_agent_form(
            agent_name=str(raw.get("agent_name","")),
            host=str(raw.get("host","")),
            port=str(raw.get("port","")),
            description=str(raw.get("description","")),
            agent_type=str(raw.get("agent_type","")),
        ).click_save()
        AssertUtil.assert_true(agent.is_agent_present(str(raw.get("agent_name",""))),
            f"Agent '{raw.get('agent_name')}' created")

    @allure.title("RDRS-1011 TC002 - Data source creation")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_rdrs_1011_tc002(self, page):
        raw = ExcelReader().get_test_data("RDRS_1011_TC002", SHEET)
        HomePage(page).open().go_to_data_source()
        ds = DataSourcePage(page)
        ds.click_add_datasource().fill_datasource_form(
            ds_name=str(raw.get("ds_name","")), db_type=str(raw.get("db_type","")),
            host=str(raw.get("host","")), port=str(raw.get("port","")),
            database=str(raw.get("database","")), username=str(raw.get("username","")),
            password=str(raw.get("password","")), schema=str(raw.get("schema","")),
            agent_name=str(raw.get("agent_name","")),
        ).click_save()
        AssertUtil.assert_true(ds.is_datasource_present(str(raw.get("ds_name",""))),
            f"DS '{raw.get('ds_name')}' created")

    @allure.title("RDRS-1011 TC003 - Process creation and execution")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_rdrs_1011_tc003(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("RDRS_1011_TC003", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_add_process().fill_process_form(
            process_name=data.process_name, process_type=data.process_type,
            source_ds=data.source_datasource, target_ds=data.target_datasource,
            source_schema=data.source_schema, source_table=data.source_table,
            target_schema=data.target_schema, target_table=data.target_table,
            transfer_mode=data.transfer_mode, truncate_target=data.truncate_target,
        ).click_save()
        HomePage(page).go_to_process_list()
        pl = ProcessListPage(page)
        pl.click_run_process(data.process_name)
        pl.wait_for_process_status(data.process_name, data.expected_status, 120000)
        AssertUtil.assert_equals(pl.get_process_status(data.process_name).upper(),
            data.expected_status.upper(), "Process status")
""")

w("tests/release_27_06_25/test_rdrs_release_27_06_25.py", """
import allure, pytest
from models.process_data import ProcessData
from pages.home_page import HomePage
from pages.process_definition_page import ProcessDefinitionPage
from pages.process_list_page import ProcessListPage
from utils.assert_util import AssertUtil
from utils.excel_reader import ExcelReader

SHEET = "Release_27_06_25"

@allure.feature("Release 27-06-25")
@pytest.mark.release_27_06_25
class TestRelease27_06_25:

    @allure.title("TC001 - Create process for release 27-06-25")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_release_tc001(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC001", SHEET))
        HomePage(page).open().go_to_process_definition()
        proc = ProcessDefinitionPage(page)
        proc.click_add_process().fill_process_form(
            process_name=data.process_name, process_type=data.process_type,
            source_ds=data.source_datasource, target_ds=data.target_datasource,
            source_schema=data.source_schema, source_table=data.source_table,
            target_schema=data.target_schema, target_table=data.target_table,
            transfer_mode=data.transfer_mode, truncate_target=data.truncate_target,
            create_table=data.create_table,
        ).click_save()
        AssertUtil.assert_true(proc.is_process_present(data.process_name),
            f"Process '{data.process_name}' visible after save")

    @allure.title("TC002 - Run process and verify completion 27-06-25")
    @allure.severity(allure.severity_level.CRITICAL)
    def test_release_tc002(self, page):
        data = ProcessData.from_dict(ExcelReader().get_test_data("TC002", SHEET))
        HomePage(page).open().go_to_process_list()
        pl = ProcessListPage(page)
        pl.click_run_process(data.process_name)
        pl.wait_for_process_status(data.process_name, data.expected_status, 120000)
        AssertUtil.assert_equals(pl.get_process_status(data.process_name).upper(),
            data.expected_status.upper(), "Process status")
""")

# ── create_excel_template.py ──────────────────────────────────────────────────
w("create_excel_template.py", """
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

HEADERS_PROCESS = [
    "test_case_id","process_name","description","process_type","transfer_mode",
    "source_datasource","source_schema","source_table",
    "target_datasource","target_schema","target_table",
    "truncate_target","create_table","schedule_enabled","cron_expression",
    "expected_status","expected_row_count",
]
HEADERS_RELEASE = HEADERS_PROCESS + [
    "agent_name","host","port","agent_type","ds_name","db_type","database","username","password","schema"
]
SHEETS = {
    "BulkTransfer_DB2_MSSQL": HEADERS_PROCESS,
    "BulkTransfer_DB2_Postgres": HEADERS_PROCESS,
    "BulkTransfer_VSAM_MSSQL": HEADERS_PROCESS,
    "Release_30_05_25": HEADERS_RELEASE,
    "Release_27_06_25": HEADERS_PROCESS,
}
SAMPLE = {h: "" for h in HEADERS_RELEASE}
SAMPLE.update({"test_case_id":"TC001","process_name":"Sample_Process","process_type":"BULK_TRANSFER",
               "transfer_mode":"FULL","expected_status":"COMPLETED","truncate_target":"true","create_table":"false"})

def style(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(str(cell.value or ""))+4, 18)

def main():
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "RDRS.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for name, headers in SHEETS.items():
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        ws.append([SAMPLE.get(h,"") for h in headers])
        style(ws); ws.freeze_panes = "A2"
    wb.save(out)
    print(f"Excel template: {out}")

if __name__ == "__main__":
    main()
""")

print("ALL FILES CREATED SUCCESSFULLY")