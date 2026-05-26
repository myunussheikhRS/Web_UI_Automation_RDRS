import os, textwrap

BASE = r"C:\Users\msheikh\RDRS_Web_Automation"

def w(rel, content):
    p = os.path.join(BASE, rel.replace("/", os.sep))
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        f.write(textwrap.dedent(content).lstrip("\n"))
    print("OK:", rel)

# ── utils/config_manager.py ──────────────────────────────────────────────────
w("utils/config_manager.py", """
import configparser, os

class ConfigManager:
    _instance = None
    _config = None
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._load()
        return cls._instance
    def _load(self):
        self._config = configparser.ConfigParser()
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        path = os.path.join(base_dir, "config", "config.ini")
        if not os.path.exists(path):
            raise FileNotFoundError(f"config.ini not found: {path}")
        self._config.read(path)
    def get(self, key, section="settings"):
        return self._config.get(section, key)
    @property
    def web_url(self): return self.get("web_url")
    @property
    def browser(self): return self.get("browser")
    @property
    def excel_path(self):
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base, self.get("excel_path"))
    @property
    def headless(self): return self.get("headless").lower() == "true"
    @property
    def implicit_wait(self): return int(self.get("implicit_wait"))
    @property
    def explicit_wait(self): return int(self.get("explicit_wait"))
    @property
    def screenshot_dir(self):
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base, self.get("screenshot_dir"))
""")

# ── utils/locator_reader.py ───────────────────────────────────────────────────
w("utils/locator_reader.py", """
import json, os

class LocatorReader:
    _instance = None
    _cache = {}
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    def _load_page(self, page_name):
        if page_name not in self._cache:
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            path = os.path.join(base, "locators", f"{page_name}.json")
            if not os.path.exists(path):
                raise FileNotFoundError(f"Locator file not found: {path}")
            with open(path, "r", encoding="utf-8") as f:
                self._cache[page_name] = json.load(f)
        return self._cache[page_name]
    def get_locator(self, page_name, key):
        data = self._load_page(page_name)
        if key not in data:
            raise KeyError(f"Locator key '{key}' not found in '{page_name}.json'")
        return data[key]
    def get_selector(self, page_name, key):
        loc = self.get_locator(page_name, key)
        t = loc["type"].lower()
        v = loc["value"]
        if t == "css": return v
        if t == "xpath": return f"xpath={v}"
        if t == "id": return f"#{v}"
        if t == "text": return f"text={v}"
        return v
""")

# ── utils/excel_reader.py ─────────────────────────────────────────────────────
w("utils/excel_reader.py", """
import os
from typing import Any, Dict, List, Optional
import openpyxl
from utils.config_manager import ConfigManager

class ExcelReader:
    _instance = None
    _workbook = None
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._load()
        return cls._instance
    def _load(self):
        path = ConfigManager().excel_path
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel not found: {path}")
        self._workbook = openpyxl.load_workbook(path, data_only=True)
    def get_all_rows(self, sheet_name):
        if sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        sheet = self._workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [{headers[i]: row[i] for i in range(len(headers))} for row in rows[1:]]
    def get_test_data(self, test_case_id, sheet_name, id_column="test_case_id"):
        for row in self.get_all_rows(sheet_name):
            if str(row.get(id_column, "")).strip() == str(test_case_id).strip():
                return row
        raise ValueError(f"Test case '{test_case_id}' not in sheet '{sheet_name}'")
    def get_multiple_rows(self, test_case_id, sheet_name, id_column="test_case_id"):
        return [r for r in self.get_all_rows(sheet_name)
                if str(r.get(id_column,"")).strip() == str(test_case_id).strip()]
""")

# ── utils/assert_util.py ──────────────────────────────────────────────────────
w("utils/assert_util.py", """
import allure, pytest

class AssertUtil:
    @staticmethod
    def assert_equals(actual, expected, field_name="value"):
        with allure.step(f"Assert '{field_name}': expected='{expected}', actual='{actual}'"):
            if actual == expected:
                allure.attach(f"PASS | {field_name}: '{actual}' == '{expected}'",
                    name=f"PASS - {field_name}", attachment_type=allure.attachment_type.TEXT)
            else:
                allure.attach(f"FAIL | {field_name}: expected='{expected}' got='{actual}'",
                    name=f"FAIL - {field_name}", attachment_type=allure.attachment_type.TEXT)
                pytest.fail(f"Assert failed '{field_name}': expected='{expected}', actual='{actual}'")
    @staticmethod
    def assert_true(condition, message="condition"):
        with allure.step(f"Assert True: {message}"):
            if condition:
                allure.attach(f"PASS | {message}", name=f"PASS - {message}",
                    attachment_type=allure.attachment_type.TEXT)
            else:
                allure.attach(f"FAIL | {message} is False", name=f"FAIL - {message}",
                    attachment_type=allure.attachment_type.TEXT)
                pytest.fail(f"Assert failed: '{message}' is not True")
    @staticmethod
    def assert_false(condition, message="condition"):
        AssertUtil.assert_true(not condition, message)
    @staticmethod
    def assert_contains(container, substring, field_name="text"):
        with allure.step(f"Assert '{field_name}' contains '{substring}'"):
            if substring in container:
                allure.attach(f"PASS | contains '{substring}'", name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT)
            else:
                allure.attach(f"FAIL | does not contain '{substring}'",
                    name=f"FAIL - {field_name}", attachment_type=allure.attachment_type.TEXT)
                pytest.fail(f"'{field_name}' value '{container}' does not contain '{substring}'")
    @staticmethod
    def assert_not_empty(value, field_name="value"):
        with allure.step(f"Assert '{field_name}' is not empty"):
            if value is not None and str(value).strip():
                allure.attach(f"PASS | '{field_name}' = '{value}'",
                    name=f"PASS - {field_name} not empty", attachment_type=allure.attachment_type.TEXT)
            else:
                allure.attach(f"FAIL | '{field_name}' is empty",
                    name=f"FAIL - {field_name} empty", attachment_type=allure.attachment_type.TEXT)
                pytest.fail(f"'{field_name}' is empty or None")
""")

# ── utils/screenshot_util.py ──────────────────────────────────────────────────
w("utils/screenshot_util.py", """
import os
from datetime import datetime
import allure
from playwright.sync_api import Page
from utils.config_manager import ConfigManager

class ScreenshotUtil:
    @staticmethod
    def capture(page: Page, name: str = "screenshot") -> str:
        d = ConfigManager().screenshot_dir
        os.makedirs(d, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in name)
        fp = os.path.join(d, f"{safe}_{ts}.png")
        page.screenshot(path=fp, full_page=True)
        with open(fp, "rb") as f:
            allure.attach(f.read(), name=name, attachment_type=allure.attachment_type.PNG)
        return fp
    @staticmethod
    def capture_on_failure(page: Page, test_name: str) -> str:
        return ScreenshotUtil.capture(page, f"FAILURE_{test_name}")
""")

w("utils/__init__.py", "")

print("utils/ done")