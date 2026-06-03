from __future__ import annotations
import time, allure
import openpyxl
from playwright.sync_api import Page
from pages.base_page import BasePage
from utils.config_manager import ConfigManager

_P = "process_list_page"

class ProcessListPage(BasePage):
    def __init__(self, page: Page): super().__init__(page)

    def _find_excel_column(self, header_index, candidates):
        for column_name in candidates:
            if column_name in header_index:
                return column_name
        return None

    def _save_workbook_with_retry(self, workbook, workbook_path, retries=5, delay_ms=400):
        last_error = None
        for attempt in range(1, retries + 1):
            try:
                workbook.save(workbook_path)
                return
            except PermissionError as exc:
                last_error = exc
                if attempt < retries:
                    self.page.wait_for_timeout(delay_ms)

        raise PermissionError(
            f"Permission denied while saving '{workbook_path}'. "
            "The Excel file is likely open or locked by another process. "
            "Close the file and retry the test."
        ) from last_error

    def _save_and_close_workbook(self, workbook, workbook_path):
        try:
            self._save_workbook_with_retry(workbook, workbook_path)
        finally:
            workbook.close()

    def _get_process_part_input_values(self):
        field_map = {
            "agent_name": "agent_name",
            "process_id": "process_id",
            "type_of_process": "type_of_process",
            "start_of_process": "start_of_process",
            "process_name_on_agent": "process_name_on_agent",
            "start_of_process_part": "start_of_process_part",
            "state_of_process_part": "state_of_process_part",
            "return_code": "return_code",
            "error_code": "error_code",
            "message": "message",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_input_reslute_values(self):
        field_map = {
            "control_record_in": "control_record_in",
            "data_block_in": "data_block_in",
            "number_of_records_read": "number_of_records_read",
            "control_record_out": "control_record_out",
            "data_block_out": "data_block_out",
            "number_of_records_written": "number_of_records_written",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_processing_applying_values(self):
        field_map = {
            "agent_name_PA": "agent_name_PA",
            "process_id_PA": "process_id_PA",
            "type_of_process_PA": "type_of_process_PA",
            "state_of_process_part_PA": "state_of_process_part_PA",
            "process_name_on_agent_PA": "process_name_on_agent_PA",
            "return_code_PA": "return_code_PA",
            "error_code_PA": "error_code_PA",
            "message_PA": "message_PA",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_processing_values(self):
        field_map = {
            "agent_name_P": "agent_name_P",
            "process_id_P": "process_id_P",
            "type_of_process_P": "type_of_process_P",
            "process_name_on_agent_P": "process_name_on_agent_P",
            "return_code_P": "return_code_P",
            "error_code_P": "error_code_P",
            "message_P": "message_P",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_processing_applying_reslute_values(self):
        field_map = {
            "control_record_in_PA": "control_record_in_PA",
            "data_block_in_PA": "data_block_in_PA",
            "number_of_records_read_PA": "number_of_records_read_PA",
            "control_record_out_PA": "control_record_out_PA",
            "data_block_out_PA": "data_block_out_PA",
            "number_of_records_written_PA": "number_of_records_written_PA",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_applying_reslute_values(self):
        field_map = {
            "control_record_in_A": "control_record_in_A",
            "data_block_in_A": "data_block_in_A",
            "number_of_records_read_A": "number_of_records_read_A",
            "control_record_out_A": "control_record_out_A",
            "data_block_out_A": "data_block_out_A",
            "number_of_records_written_A": "number_of_records_written_A",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_applying_values(self):
        field_map = {
            "agent_name_A": "agent_name_A",
            "process_id_A": "process_id_A",
            "type_of_process_A": "type_of_process_A",
            "process_name_on_agent_A": "process_name_on_agent_A",
            "return_code_A": "return_code_A",
            "error_code_A": "error_code_A",
            "message_A": "message_A",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    def _get_process_part_processing_reslute_values(self):
        field_map = {
            "control_record_in_P": "control_record_in_P",
            "data_block_in_P": "data_block_in_P",
            "number_of_records_read_P": "number_of_records_read_P",
            "control_record_out_P": "control_record_out_P",
            "data_block_out_P": "data_block_out_P",
            "number_of_records_written_P": "number_of_records_written_P",
        }
        return {
            field_name: self.get_text(self._loc(_P, locator_key).first).strip()
            for field_name, locator_key in field_map.items()
        }

    # written by Yunus
    @allure.step("Click processList: {process_name}")
    def click_process_list(self, process_name):
        self.page.get_by_text("Process definitions", exact=True).first.click()
        self.page.keyboard.press("ArrowLeft")

        pl = self.page.get_by_text("Process list", exact=True).first
        pl.wait_for(state="visible", timeout=self._timeout)

        # First open attempt
        try:
            pl.click(timeout=1500)
        except Exception:
            pl.click(force=True, timeout=1500)

        active_process = self._loc(_P, "active_process_list")
        deadline = time.time() + self._timeout / 1000

        # While active panel is visible, keep toggling Process list with short retries
        while time.time() < deadline and active_process.is_visible():
            try:
                pl.click(timeout=1000)
            except Exception:
                pl.click(force=True, timeout=1000)
            self.page.wait_for_timeout(250)

        if active_process.is_visible():
            raise TimeoutError("active_process_list stayed visible after retries")

        loc = self.page.locator(f"(//div[contains(text(),'{process_name}')])[1]")
        loc.wait_for(state="visible", timeout=self._timeout)
        loc.click()
        return self


    @allure.step("Validate process part input values")
    def validate_process_part_input(self, expected_values):
        # POM should only validate UI state; expected values are provided by the test.
        field_map = {
            "agent_name": "agent_name",
            "process_id": "process_id",
            "type_of_process": "type_of_process",
            "start_of_process": "start_of_process",
            "process_name_on_agent": "process_name_on_agent",
            "start_of_process_part": "start_of_process_part",
            "state_of_process_part": "state_of_process_part",
            "return_code": "return_code",
            "error_code": "error_code",
            "message": "message",
        }

        # Accept both internal snake_case keys and Excel-style keys.
        expected_key_aliases = {
            "agent_name": ["agent_name", "Agent"],
            "process_id": ["process_id", "Process_ID", "Process_Id"],
            "type_of_process": ["type_of_process", "Type_Of_Process"],
            "start_of_process": ["start_of_process", "Start_Of_Process"],
            "process_name_on_agent": ["process_name_on_agent", "Process_Name_On_Agent"],
            "start_of_process_part": ["start_of_process_part", "Start_Of_Process_Part"],
            "state_of_process_part": ["state_of_process_part", "State_Of_Process_Part"],
            "return_code": ["return_code", "Return_Code"],
            "error_code": ["error_code", "Error_Code"],
            "message": ["message", "Message"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"agent_name": expected_values}

        actual_values = self._get_process_part_input_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            # Skip empty expectations so test data can validate only relevant fields.
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                # Short PASS info
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                # Detailed FAIL info
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected process-part fields were provided. "
                "Pass values using keys like agent_name/Agent, process_id/Process_ID, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_input_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_input_values()
        excel_field_map = {
            "agent_name": ["Agent"],
            "process_id": ["Process_ID"],
            "type_of_process": ["Type_Of_Process"],
            "start_of_process": ["Start_Of_Process"],
            "process_name_on_agent": ["Process_Name_On_Agent"],
            "start_of_process_part": ["Start_Of_Process_Part"],
            "state_of_process_part": ["State_Of_Process_Part"],
            "return_code": ["Return_Code"],
            "error_code": ["Error_Code"],
            "message": ["Message"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for process part values"
            )

        return self

    @allure.step("Validate process part input reslute values")
    def validate_process_part_input_reslute(self, expected_values):
        field_map = {
            "control_record_in": "control_record_in",
            "data_block_in": "data_block_in",
            "number_of_records_read": "number_of_records_read",
            "control_record_out": "control_record_out",
            "data_block_out": "data_block_out",
            "number_of_records_written": "number_of_records_written",
        }

        expected_key_aliases = {
            "control_record_in": ["control_record_in", "Control_Record_In"],
            "data_block_in": ["data_block_in", "Data_Block_In"],
            "number_of_records_read": ["number_of_records_read", "Number_Of_Records_Read"],
            "control_record_out": ["control_record_out", "Control_Record_Out"],
            "data_block_out": ["data_block_out", "Data_Block_Out"],
            "number_of_records_written": ["number_of_records_written", "Number_Of_Records_Written"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"control_record_in": expected_values}

        actual_values = self._get_process_part_input_reslute_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected process-part-reslute fields were provided. "
                "Pass values using keys like control_record_in/Control_Record_In, "
                "number_of_records_read/Number_Of_Records_Read, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_input_reslute_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Input Reslute Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part input reslute validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part input reslute values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_input_reslute_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_input_reslute_values()
        excel_field_map = {
            "control_record_in": ["Control_Record_In", "Control_Records_In"],
            "data_block_in": ["Data_Block_In"],
            "number_of_records_read": ["Number_Of_Records_Read", "Number_Of_Recodrs_Read"],
            "control_record_out": ["Control_Record_Out", "Control_Records_Out"],
            "data_block_out": ["Data_Block_Out"],
            "number_of_records_written": ["Number_Of_Records_Written"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for process part input reslute values"
            )

        return self

    @allure.step("Validate process part Processing values")
    def validate_process_part_Processing(self, expected_values):
        field_map = {
            "agent_name_P": "agent_name_P",
            "process_id_P": "process_id_P",
            "type_of_process_P": "type_of_process_P",
            "process_name_on_agent_P": "process_name_on_agent_P",
            "return_code_P": "return_code_P",
            "error_code_P": "error_code_P",
            "message_P": "message_P",
        }

        expected_key_aliases = {
            "agent_name_P": ["agent_name_P", "Agent_P"],
            "process_id_P": ["process_id_P", "Process_ID_P", "Process_Id_P"],
            "type_of_process_P": ["type_of_process_P", "Type_Of_Process_P"],
            "process_name_on_agent_P": ["process_name_on_agent_P", "Process_Name_On_Agent_P"],
            "return_code_P": ["return_code_P", "Return_Code_P"],
            "error_code_P": ["error_code_P", "Error_Code_P"],
            "message_P": ["message_P", "Message_P"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"agent_name_P": expected_values}

        actual_values = self._get_process_part_processing_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected Processing fields were provided. "
                "Pass values using keys like agent_name_P/Agent_P, process_id_P/Process_ID_P, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_processing_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Processing Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part Processing validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part Processing values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_Processing_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_processing_values()
        excel_field_map = {
            "agent_name_P": ["Agent_P", "AgentName_P"],
            "process_id_P": ["Process_ID_P", "Process_Id_P"],
            "type_of_process_P": ["Type_Of_Process_P"],
            "process_name_on_agent_P": ["Process_Name_On_Agent_P"],
            "return_code_P": ["Return_Code_P"],
            "error_code_P": ["Error_Code_P"],
            "message_P": ["Message_P"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for Processing values"
            )

        return self

    @allure.step("Validate process part ProcessingApplying values")
    def validate_process_part_ProcessingApplying(self, expected_values):
        field_map = {
            "agent_name_PA": "agent_name_PA",
            "process_id_PA": "process_id_PA",
            "type_of_process_PA": "type_of_process_PA",
            "state_of_process_part_PA": "state_of_process_part_PA",
            "process_name_on_agent_PA": "process_name_on_agent_PA",
            "return_code_PA": "return_code_PA",
            "error_code_PA": "error_code_PA",
            "message_PA": "message_PA",
        }

        expected_key_aliases = {
            "agent_name_PA": ["agent_name_PA", "Agent_PA"],
            "process_id_PA": ["process_id_PA", "Process_ID_PA", "Process_Id_PA"],
            "type_of_process_PA": ["type_of_process_PA", "Type_Of_Process_PA"],
            "state_of_process_part_PA": ["state_of_process_part_PA", "State_Of_Process_Part_PA"],
            "process_name_on_agent_PA": ["process_name_on_agent_PA", "Process_Name_On_Agent_PA"],
            "return_code_PA": ["return_code_PA", "Return_Code_PA"],
            "error_code_PA": ["error_code_PA", "Error_Code_PA"],
            "message_PA": ["message_PA", "Message_PA"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"agent_name_PA": expected_values}

        actual_values = self._get_process_part_processing_applying_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected ProcessingApplying fields were provided. "
                "Pass values using keys like agent_name_PA/Agent_PA, process_id_PA/Process_ID_PA, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_processing_applying_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part ProcessingApplying Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part ProcessingApplying validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part ProcessingApplying values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_processingApplying_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_processing_applying_values()
        excel_field_map = {
            "agent_name_PA": ["Agent_PA", "AgentName_PA"],
            "process_id_PA": ["Process_ID_PA", "Process_Id_PA"],
            "type_of_process_PA": ["Type_Of_Process_PA"],
            "state_of_process_part_PA": ["State_Of_Process_Part_PA"],
            "process_name_on_agent_PA": ["Process_Name_On_Agent_PA"],
            "return_code_PA": ["Return_Code_PA"],
            "error_code_PA": ["Error_Code_PA"],
            "message_PA": ["Message_PA"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for ProcessingApplying values"
            )

        return self

    @allure.step("Validate process part ProcessingApplying reslute values")
    def validate_process_part_ProcessingApplying_reslute(self, expected_values):
        field_map = {
            "control_record_in_PA": "control_record_in_PA",
            "data_block_in_PA": "data_block_in_PA",
            "number_of_records_read_PA": "number_of_records_read_PA",
            "control_record_out_PA": "control_record_out_PA",
            "data_block_out_PA": "data_block_out_PA",
            "number_of_records_written_PA": "number_of_records_written_PA",
        }

        expected_key_aliases = {
            "control_record_in_PA": ["control_record_in_PA", "Control_Record_In_PA"],
            "data_block_in_PA": ["data_block_in_PA", "Data_Block_In_PA"],
            "number_of_records_read_PA": ["number_of_records_read_PA", "Number_Of_Records_Read_PA"],
            "control_record_out_PA": ["control_record_out_PA", "Control_Record_Out_PA"],
            "data_block_out_PA": ["data_block_out_PA", "Data_Block_Out_PA"],
            "number_of_records_written_PA": ["number_of_records_written_PA", "Number_Of_Records_Written_PA"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"control_record_in_PA": expected_values}

        actual_values = self._get_process_part_processing_applying_reslute_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected ProcessingApplying reslute fields were provided. "
                "Pass values using keys like control_record_in_PA/Control_Record_In_PA, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_processing_applying_reslute_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part ProcessingApplying Reslute Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part ProcessingApplying reslute validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part ProcessingApplying reslute values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_processingApplying_reslute_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_processing_applying_reslute_values()
        excel_field_map = {
            "control_record_in_PA": ["Control_Record_In_PA", "Control_Records_In_PA"],
            "data_block_in_PA": ["Data_Block_In_PA"],
            "number_of_records_read_PA": ["Number_Of_Records_Read_PA"],
            "control_record_out_PA": ["Control_Record_Out_PA", "Control_Records_Out_PA"],
            "data_block_out_PA": ["Data_Block_Out_PA"],
            "number_of_records_written_PA": ["Number_Of_Records_Written_PA"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for ProcessingApplying reslute values"
            )

        return self

    @allure.step("Validate process part applying reslute values")
    def validate_process_part_applying_reslute(self, expected_values):
        field_map = {
            "control_record_in_A": "control_record_in_A",
            "data_block_in_A": "data_block_in_A",
            "number_of_records_read_A": "number_of_records_read_A",
            "control_record_out_A": "control_record_out_A",
            "data_block_out_A": "data_block_out_A",
            "number_of_records_written_A": "number_of_records_written_A",
        }

        expected_key_aliases = {
            "control_record_in_A": ["control_record_in_A", "Control_Record_In_A", "Control_Records_In_A"],
            "data_block_in_A": ["data_block_in_A", "Data_Block_In_A"],
            "number_of_records_read_A": ["number_of_records_read_A", "Number_Of_Records_Read_A", "Number_Of_Recodrs_Read_A"],
            "control_record_out_A": ["control_record_out_A", "Control_Record_Out_A", "Control_Records_Out_A"],
            "data_block_out_A": ["data_block_out_A", "Data_Block_Out_A"],
            "number_of_records_written_A": ["number_of_records_written_A", "Number_Of_Records_Written_A"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"control_record_in_A": expected_values}

        actual_values = self._get_process_part_applying_reslute_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected applying-reslute fields were provided. "
                "Pass values using keys like control_record_in_A/Control_Record_In_A, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_applying_reslute_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Applying Reslute Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part applying reslute validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part applying reslute values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_applying_reslute_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_applying_reslute_values()
        excel_field_map = {
            "control_record_in_A": ["Control_Record_In_A", "Control_Records_In_A"],
            "data_block_in_A": ["Data_Block_In_A"],
            "number_of_records_read_A": ["Number_Of_Records_Read_A", "Number_Of_Recodrs_Read_A"],
            "control_record_out_A": ["Control_Record_Out_A", "Control_Records_Out_A"],
            "data_block_out_A": ["Data_Block_Out_A"],
            "number_of_records_written_A": ["Number_Of_Records_Written_A"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for applying reslute values"
            )

        return self

    @allure.step("Validate process part Applying values")
    def validate_process_part_Applying(self, expected_values):
        field_map = {
            "agent_name_A": "agent_name_A",
            "process_id_A": "process_id_A",
            "type_of_process_A": "type_of_process_A",
            "process_name_on_agent_A": "process_name_on_agent_A",
            "return_code_A": "return_code_A",
            "error_code_A": "error_code_A",
            "message_A": "message_A",
        }

        expected_key_aliases = {
            "agent_name_A": ["agent_name_A", "Agent_A"],
            "process_id_A": ["process_id_A", "Process_ID_A", "Process_Id_A"],
            "type_of_process_A": ["type_of_process_A", "Type_Of_Process_A"],
            "process_name_on_agent_A": ["process_name_on_agent_A", "Process_Name_On_Agent_A"],
            "return_code_A": ["return_code_A", "Return_Code_A"],
            "error_code_A": ["error_code_A", "Error_Code_A"],
            "message_A": ["message_A", "Message_A"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"agent_name_A": expected_values}

        actual_values = self._get_process_part_applying_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected Applying fields were provided. "
                "Pass values using keys like agent_name_A/Agent_A, process_id_A/Process_ID_A, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_applying_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Applying Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part Applying validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part Applying values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_Applying_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_applying_values()
        excel_field_map = {
            "agent_name_A": ["Agent_A"],
            "process_id_A": ["Process_ID_A", "Process_Id_A"],
            "type_of_process_A": ["Type_Of_Process_A"],
            "process_name_on_agent_A": ["Process_Name_On_Agent_A"],
            "return_code_A": ["Return_Code_A"],
            "error_code_A": ["Error_Code_A"],
            "message_A": ["Message_A"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for Applying values"
            )

        return self

    @allure.step("Validate process part Processing reslute values")
    def validate_process_part_Processing_reslute(self, expected_values):
        field_map = {
            "control_record_in_P": "control_record_in_P",
            "data_block_in_P": "data_block_in_P",
            "number_of_records_read_P": "number_of_records_read_P",
            "control_record_out_P": "control_record_out_P",
            "data_block_out_P": "data_block_out_P",
            "number_of_records_written_P": "number_of_records_written_P",
        }

        expected_key_aliases = {
            "control_record_in_P": ["control_record_in_P", "Control_Record_In_P", "Control_Records_In_P"],
            "data_block_in_P": ["data_block_in_P", "Data_Block_In_P"],
            "number_of_records_read_P": ["number_of_records_read_P", "Number_Of_Records_Read_P", "Number_Of_Recodrs_Read_P"],
            "control_record_out_P": ["control_record_out_P", "Control_Record_Out_P", "Control_Records_Out_P"],
            "data_block_out_P": ["data_block_out_P", "Data_Block_Out_P"],
            "number_of_records_written_P": ["number_of_records_written_P", "Number_Of_Records_Written_P"],
        }

        if not isinstance(expected_values, dict):
            expected_values = {"control_record_in_P": expected_values}

        actual_values = self._get_process_part_processing_reslute_values()
        mismatches = []
        validated_count = 0

        for field_name in field_map.keys():
            aliases = expected_key_aliases[field_name]
            source_key = next((k for k in aliases if k in expected_values), None)
            if source_key is None:
                continue

            raw_expected = expected_values[source_key]
            if raw_expected is None or str(raw_expected).strip() == "":
                continue

            expected_value = str(raw_expected).strip()
            actual_value = actual_values[field_name]
            validated_count += 1

            if actual_value == expected_value:
                allure.attach(
                    f"PASS | {field_name}",
                    name=f"PASS - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
            else:
                detail = (
                    f"Field: {field_name}\n"
                    f"Expected: {expected_value}\n"
                    f"Actual: {actual_value}\n"
                    f"Source key: {source_key}"
                )
                allure.attach(
                    detail,
                    name=f"FAIL - {field_name}",
                    attachment_type=allure.attachment_type.TEXT,
                )
                mismatches.append(
                    f"{field_name}: actual='{actual_value}' expected='{expected_value}'"
                )

        if validated_count == 0:
            msg = (
                "No comparable expected Processing-reslute fields were provided. "
                "Pass values using keys like control_record_in_P/Control_Record_In_P, etc."
            )
            allure.attach(msg, name="Validation Input Error", attachment_type=allure.attachment_type.TEXT)
            raise AssertionError(msg)

        if mismatches:
            self.page.screenshot(path="screenshots/process_part_processing_reslute_mismatch.png")
            allure.attach(
                "\n".join(mismatches),
                name="Process Part Processing Reslute Mismatch Summary",
                attachment_type=allure.attachment_type.TEXT,
            )
            raise AssertionError("Process part Processing reslute validation failed:\n" + "\n".join(mismatches))

        return self

    @allure.step("Save process part Processing reslute values to Excel for {test_case_id} in {sheet_name}")
    def save_process_part_Processing_reslute_to_excel(self, test_case_id, sheet_name, id_column="TestCase", part_name=None):
        ui_values = self._get_process_part_processing_reslute_values()
        excel_field_map = {
            "control_record_in_P": ["Control_Record_In_P", "Control_Records_In_P"],
            "data_block_in_P": ["Data_Block_In_P"],
            "number_of_records_read_P": ["Number_Of_Records_Read_P", "Number_Of_Recodrs_Read_P"],
            "control_record_out_P": ["Control_Record_Out_P", "Control_Records_Out_P"],
            "data_block_out_P": ["Data_Block_Out_P"],
            "number_of_records_written_P": ["Number_Of_Records_Written_P"],
        }

        workbook_path = ConfigManager().excel_path
        workbook = openpyxl.load_workbook(workbook_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{workbook_path}'")

        sheet = workbook[sheet_name]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        if id_column not in header_index:
            raise ValueError(f"Column '{id_column}' not found in sheet '{sheet_name}'")

        row_index = None
        for current_row in range(2, sheet.max_row + 1):
            value = sheet.cell(current_row, header_index[id_column]).value
            if str(value or "").strip() == str(test_case_id).strip():
                row_index = current_row
                break

        if row_index is None:
            raise ValueError(f"Test case '{test_case_id}' not found in sheet '{sheet_name}'")

        updated_fields = []
        for ui_field, excel_candidates in excel_field_map.items():
            excel_column = self._find_excel_column(header_index, excel_candidates)
            if excel_column is None:
                continue
            sheet.cell(row_index, header_index[excel_column]).value = ui_values[ui_field]
            updated_fields.append(excel_column)

        self._save_and_close_workbook(workbook, workbook_path)

        if not updated_fields:
            raise AssertionError(
                f"No matching Excel columns found in sheet '{sheet_name}' for Processing reslute values"
            )

        return self

        

   

    @allure.step("Wait for process '{process_name}' to reach '{expected_status}'")
    def wait_for_process_status(self, process_name, expected_status, timeout_ms=60000):
        deadline = time.time() + timeout_ms / 1000
        while time.time() < deadline:
            if self.get_process_status(process_name).upper() == expected_status.upper():
                return self
            self.click_refresh()
            self.page.wait_for_timeout(2000)
        raise TimeoutError(f"Process '{process_name}' did not reach '{expected_status}'")

    def get_page_title(self): return self.get_text(self._loc(_P, "page_title"))
    def get_process_status(self, process_name):
        return self._get_process_row(process_name).locator("td.process-status, td:nth-child(3), span.status").inner_text().strip()
    def is_process_present(self, process_name):
        rows = self._loc(_P, "process_table_rows")
        return any(process_name in rows.nth(i).inner_text() for i in range(rows.count()))
    def get_log_content(self): return self.get_text(self._loc(_P, "log_content"))
    def get_process_count(self): return self._loc(_P, "process_table_rows").count()
    def _get_process_row(self, process_name):
        rows = self._loc(_P, "process_table_rows")
        for i in range(rows.count()):
            r = rows.nth(i)
            if process_name in r.inner_text(): return r
        raise ValueError(f"Process not found in list: {process_name}")
