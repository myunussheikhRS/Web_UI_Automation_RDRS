import os
from typing import Any, Dict, List, Optional
import openpyxl
from utils.config_manager import ConfigManager

class ExcelReader:
    _instance = None
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def _workbook_path(self):
        path = ConfigManager().excel_path
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel not found: {path}")
        return path

    def get_all_rows(self, sheet_name):
        workbook = openpyxl.load_workbook(self._workbook_path(), data_only=True, read_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            return [{headers[i]: row[i] for i in range(len(headers))} for row in rows[1:]]
        finally:
            workbook.close()
    def get_test_data(self, test_case_id, sheet_name, id_column="test_case_id"):
        for row in self.get_all_rows(sheet_name):
            if str(row.get(id_column, "")).strip() == str(test_case_id).strip():
                return row
        raise ValueError(f"Test case '{test_case_id}' not in sheet '{sheet_name}'")
    def get_multiple_rows(self, test_case_id, sheet_name, id_column="test_case_id"):
        return [r for r in self.get_all_rows(sheet_name)
                if str(r.get(id_column,"")).strip() == str(test_case_id).strip()]
